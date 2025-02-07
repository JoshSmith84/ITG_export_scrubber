"""
ITG_export_scrub_gui.py

Author: Josh Smith

Purpose: Stage ITG export data into a single workbook per client with
separate worksheets based on input data csv's. Format cells based on
width of widest string, freeze header row/format as a table,
and zip for sending. Also edit cell data to remove html, unicode issues,
and leave only readable data.
See 'TPG ITG Export Scrubber Specification.rst' for more info.
"""

# imports
import os
from os.path import basename
import shutil
import csv
import unicodedata
from bs4 import BeautifulSoup
from zipfile36 import ZipFile
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import sys
import logging
import datetime
import pandas as pd
import xlsxwriter
import traceback


class LabelInput(tk.Frame):
    """A widget containing a label and input together.
    Credit: Alan D. Moore "Python GUI Programming with Tkinter"""

    def __init__(
            self, parent, label, var, input_class=ttk.Entry,
            input_args=None, label_args=None, **kwargs
    ):
        super().__init__(parent, **kwargs)
        input_args = input_args or {}
        label_args = label_args or {}
        self.variable = var
        self.variable.label_widget = self

        if input_class in (ttk.Checkbutton, ttk.Button):
            input_args["text"] = label
        else:
            self.label = ttk.Label(self, text=label, **label_args)
            self.label.grid(row=0, column=0, sticky=(tk.W + tk.E))

        if input_class in (
            ttk.Checkbutton, ttk.Button, ttk.Radiobutton
        ):
            input_args["variable"] = self.variable
        else:
            input_args["textvariable"] = self.variable

        if input_class == ttk.Radiobutton:
            self.input = tk.Frame(self)
            for v in input_args.pop('values', []):
                button = ttk.Radiobutton(
                    self.input, value=v, text=v, **input_args
                )
                button.pack(
                    side=tk.LEFT, ipadx=10,
                    ipady=2, expand=True, fill='x'
                )
        else:
            self.input = input_class(self, **input_args)

        self.input.grid(row=1, column=0, sticky=(tk.E + tk.W))
        self.columnconfigure(0, weight=1)

    def grid(self, sticky=(tk.E + tk.W), **kwargs):
        """Override grid to add default sticky values"""
        super().grid(sticky=sticky, **kwargs)


class AppPage(ttk.Frame):
    """Application page class from which all other pages will inherit."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._vars = {}

    def _add_frame(self, label, cols=2):
        """Add a label frame to the form
        Credit: Alan D. Moore "Python GUI Programming with Tkinter"""

        frame = ttk.LabelFrame(self, text=label)
        frame.grid(sticky=tk.W + tk.E)
        for i in range(cols):
            frame.columnconfigure(i, weight=1)
        return frame

    def get(self):
        data = dict()
        for key, variable in self._vars.items():
            try:
                data[key] = variable.get()
            except tk.TclError:
                message = f'Error in field: {key}.'
                raise ValueError(message)
        return data


class MainPage(AppPage):
    """Main Page to select options, change folder, and run"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._vars = {'Batch Size': tk.StringVar(None, 'Folder'),
                      'Post Job': tk.StringVar(None, 'Delete'),
                      'Zip?': tk.StringVar(None, 'No'),
                      }

        self.input_folder = ''
        self.input_file = ''
        self.err_present = 0
        self.err_count = 0

        # Initialize Main Page GUI
        size_default = self._add_frame(
            'Processing a folder or single file?'
        )
        post_default = self._add_frame(
            'Delete or Keep original export when finished?'
        )
        zip_default = self._add_frame('Zip the output when finished?')
        buttons = self._add_frame('')

        LabelInput(size_default, '', input_class=ttk.Radiobutton,
                   var=self._vars['Batch Size'],
                   input_args={'values': ['Folder', 'Single File']}
                   ).grid(row=0, column=0, sticky=(tk.W + tk.E)
                          )

        LabelInput(post_default, '', input_class=ttk.Radiobutton,
                   var=self._vars['Post Job'],
                   input_args={'values': ['Delete', 'Keep']}
                   ).grid(row=1, column=0, sticky=(tk.W + tk.E)
                          )

        LabelInput(zip_default, '', input_class=ttk.Radiobutton,
                   var=self._vars['Zip?'],
                   input_args={'values': ['Yes', 'No']}
                   ).grid(row=2, column=0, sticky=(tk.W + tk.E)
                          )

        self.run_button = tk.Button(buttons, text='Run',
                                    command=self._on_run
                                    )
        self.run_button.grid(row=0, column=1, sticky='ew')

        self.select_target = tk.Button(buttons, text='Select Target',
                                       command=self._on_target
                                       )
        self.select_target.grid(row=0, column=0, sticky='ew')

        self.quit_button = tk.Button(
            buttons,
            text='Quit',
            command=self._on_quit
        )
        self.quit_button.grid(row=3, column=0, sticky='ew')

        self.status = tk.StringVar(
            None, 'Status: '
                  'Please select a target to continue...'
        )
        ttk.Label(
            self, textvariable=self.status, wraplength=225, justify='left'
        ).grid(sticky=(tk.W + tk.E), row=4, padx=10)

    def process_exports(self, input_zip, post_task, zip_task) -> int:
        """Main processing method. Take a TPG ITG export, unzip it,
        ignore unneeded data, clean left over csv's of empty columns,
        and make it readable

        :param input_zip: any zip file. Non ITG exports will be unzipped,
                        and ignored once contents are detected as invalid.
        :param post_task: Option to either delete or keep input zip
                        when processing is complete.
        :param zip_task: Option to either zip the output or not
                        when processing is complete.
        :return: Integer 0 or 1. 0 means no error occurred.
        1 means error log is present.
        """

        # List of CSV's to extract. Ignore all else
        keep_csv = [
            'applications-licensing.csv', 'backup.csv',
            'backups-managed.csv',
            'battery-backup-ups.csv', 'configurations.csv',
            'domain-hosting.csv',
            'email.csv', 'file-sharing.csv', 'internet-wan.csv',
            'lan.csv', 'passwords.csv', 'printing.csv',
            'vendors.csv', 'voice-pbx-fax.csv', 'wireless.csv',
                    ]

        # Column names that if left in first column after processing,
        # should be sorted by in the final sheet
        sorters = ['Name', 'name', 'Hostname', 'Printer Name',
                   'Description', 'Vendor Name'
                   ]

        # Prep the processing paths
        explode_path = input_zip.split('/')
        new_explode_path = [i.replace('/', '\\') for i in explode_path]
        input_zip = '\\'.join(new_explode_path)
        new_explode_path.pop(-1)
        working_dir = '\\'.join(new_explode_path) + '\\'
        export_dir = working_dir + 'itg_unzipped\\'
        error_log = working_dir + (f'ITG_scrubber_errors_'
                                   f'{datetime.date.today()}.txt')

        # Dictionary to contain all customer sheets:dataframe
        sheets_dict = {}

        # Unzip input
        try:
            with ZipFile(input_zip, 'r') as in_zip:
                for file in in_zip.infolist():
                    if file.filename in keep_csv:
                        in_zip.extract(file, export_dir)
        except FileNotFoundError:
            self.log_error(error_log, f'{input_zip} not found. '
                                      f'More Info: {traceback.format_exc()}'
                                      f'\n\n'
                           )
            return 1
        except PermissionError:
            self.log_error(error_log, f'{input_zip} '
                                      f'permission denied.'
                                      f' Try Running again as admin. '
                                      f'More Info: {traceback.format_exc()}'
                                      f'\n\n'
                           )
            return 1
        except in_zip.BadZipFile:
            self.log_error(error_log, f'{input_zip}'
                                      f' may be corrupt. '
                                      f'More Info: {traceback.format_exc()}'
                                      f'\n\n'
                           )
            return 1
        except OSError:
            self.log_error(error_log, f'{input_zip} '
                                      f'caused an OS error. '
                                      f' Drive may be full or '
                                      f'path is no longer valid. '
                                      f'More Info: {traceback.format_exc()}'
                                      f'\n\n'
                           )
            return 1

        # Check if the zip is a valid export(export_dir will not exist)
        if os.path.exists(export_dir):
            # Gather list of files
            raw_input_files = [f for f in os.listdir(export_dir)]
        else:
            self.status.set(f'{input_zip} is not a valid ITG export')
            Application.update(self)
            return 0

        # If Backup-managed is present, ignore backup
        if 'backups-managed.csv' in raw_input_files:
            input_files = [f for f in raw_input_files if f != 'backup.csv']
        else:
            input_files = [f for f in raw_input_files]

        # From any of the csvs, pull customer name from column B
        with open(export_dir + input_files[0], 'r',
                  encoding='utf-8') as csv_file:
            headers = csv_file.readline().strip('\n').split(',')
            reader = csv.reader(csv_file)
            customer_name = list(reader)[0][1]

        # Iterate through every remaining csv,
        # and make changes in memory
        for file in input_files:
            self.status.set(f'Processing {customer_name} ...')
            Application.update(self)
            # Reset columns to delete list
            # (columns that always are deleted/ignored)
            delete_columns = ['id', 'organization', 'Category',
                              'Business Impact',
                              'Client Subject Matter Expert',
                              'Importance', 'archived',
                              'Backup Estimated Start Date',
                              'FlexAssset Review Date',
                              'FlexAsset Review Date',
                              'Backup Radar Reporting Schedule',
                              'hostname', 'manufacturer',
                              'position', 'contact', 'location',
                              'configuration_interfaces',
                              'DHCP Exclusions', 'one_time_password',
                              'Printer Management Login',
                              'installed_by',
                              'Equipment make & Model',
                              'resource_type', 'resource_id',
                              'configuration_status', 'asset_tag',
                              'DHCP Server', 'DHCP Scope',
                              'DHCP Reservations', 'DNS Server(s)',
                              'Default Gateway Device', 'Firewall',
                              'Access Point(s)',
                              'Wireless Controller (Application)',
                              'Wireless Controller (Hardware)',
                              'Management Credentials', 'VLAN #',
                              'Backup Radar Report Recipients (Email)'
                              ' or Link Contacts',
                              'Backup Radar Reporting Notes',
                              'Backup Server/NAS Management Login',
                              'Local Backup Encryption Key',
                              'Backup Copy Job Name',
                              'Backup Copy Target',
                              'Backup Copy Encryption',
                              'Configuration Backup to Cloud Connect?',
                              'SMB Login',
                              ]

            # Continue with unpacking current csv to list of lists
            working_rows = []
            with open(export_dir + file, 'r', encoding='utf-8') as csv_file:
                headers = csv_file.readline().strip('\n').split(',')
                reader = csv.reader(csv_file)
                for row in reader:
                    new_row = []
                    for cell in row:
                        # Detect html in cell and convert if so.
                        if bool(
                                BeautifulSoup(cell, 'lxml'
                                              ).find()
                                ):
                            cell = BeautifulSoup(cell, 'lxml').text

                        # normalize text
                        cell = unicodedata.normalize(
                            'NFKD', cell
                        )
                        new_row.append(cell)
                    working_rows.append(new_row)

            # Find the archive column and keep track of it
            # (it's usually last but not always)
            # Also find config status column
            archive_index = -1
            configuration_status_index = 0
            for index, value in enumerate(headers):
                if value == 'archived':
                    archive_index = index
                if file == 'configurations.csv':
                    if value == 'configuration_status':
                        configuration_status_index = index

            # go through every row and delete any row with archive set to 'Yes'
            # and any configuration status in configurations csv
            # other than Active
            top_index_current = len(working_rows) - 1
            for index, value in enumerate(reversed(working_rows)):
                if value[archive_index] == 'Yes':
                    del working_rows[top_index_current - index]
                if configuration_status_index != 0:
                    if value[configuration_status_index] != 'Active':
                        del working_rows[top_index_current - index]

            # Find empty columns
            for i in range(len(headers)):
                i_empty = 0
                for value in working_rows:
                    if value[i] == '':
                        i_empty += 1
                if i_empty == len(working_rows):
                    delete_columns.append(headers[i])

            # Ignore all blank columns and columns to be "deleted"
            clean_rows = []
            for row in working_rows:
                new_row = []
                for i in range(len(row)):
                    if headers[i] not in delete_columns:
                        new_row.append(row[i])
                clean_rows.append(new_row)

            # Clean up headers to match
            new_headers = [h for h in headers if h not in delete_columns]

            # Convert the list of lists to a pandas DataFrame
            columns_dict = {}
            for row in clean_rows:
                for i, value in enumerate(row):
                    columns_dict.setdefault(new_headers[i], []).append(value)
            df = pd.DataFrame(columns_dict)

            # Sort any sheets if name column is present and
            # Populate the Pandas dataframe into a dictionary to
            # populate the workbook after all data is processed
            if len(new_headers) > 0:
                if new_headers[0] in sorters:
                    sheets_dict.update({file.split('.')[0]: df.sort_values(
                        by=new_headers[0])}
                    )
                else:
                    sheets_dict.update({file.split('.')[0]: df})
            else:
                continue

        # Populate Workbook with all sheets/tables and data
        wb_file = working_dir + f'{customer_name}_export.xlsx'
        if os.path.exists(wb_file):
            try:
                os.remove(wb_file)
            except PermissionError:
                self.log_error(error_log, f'Attempted '
                                          f'deleting old {wb_file}, '
                                          f' but permission denied.'
                                          f' Try running again as admin. '
                                          f'More Info: {traceback.format_exc()}'
                                          f'\n\n'
                               )
                shutil.rmtree(export_dir)
                return 1
        wb = xlsxwriter.Workbook(wb_file)
        for key, value in sheets_dict.items():
            sheet = wb.add_worksheet(key)
            sheet.add_table(0, 0, value.shape[0], value.shape[1] - 1, {
                'data': value.values.tolist(),
                'columns': [{'header': col} for col in value.columns]
            })
        wb.close()

        # Delete the unzipped original export
        shutil.rmtree(export_dir)

        # Open workbook again with openpyxl and make final adjustments
        # Find and set uniform column width
        wb = load_workbook(wb_file)
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                if max_length > 50:
                    max_length = 50
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width

        wb.save(wb_file)

        # Delete or keep unzipped export
        if post_task == 'Delete':
            try:
                os.remove(input_zip)
            except PermissionError:
                self.log_error(error_log, f'Attempted '
                                          f'deleting {input_zip}, '
                                          f' but permission denied.'
                                          f' Try running again as admin. '
                                          f'More Info: {traceback.format_exc()}'
                                          f'\n\n'
                               )
                return 1

        self.status.set(f'Processing of {customer_name} complete.')
        Application.update(self)

        # To zip or not to zip output file (needs to be zipped for email)
        if zip_task == 'Yes':
            out_zip = f'{working_dir}\\{customer_name}_export.zip'
            if os.path.exists(out_zip):
                try:
                    os.remove(out_zip)
                except PermissionError:
                    self.log_error(error_log, f'Attempted '
                                              f'deleting {input_zip}, '
                                              f' but permission denied.'
                                              f' Try running again as admin. '
                                              f'More Info: '
                                              f'{traceback.format_exc()}'
                                              f'\n\n'
                                   )
                    return 1
            with ZipFile(out_zip, 'w') as f:
                f.write(wb_file, basename(wb_file))
            try:
                os.remove(wb_file)
            except PermissionError:
                self.log_error(error_log, f'Attempted '
                                          f'deleting {wb_file} '
                                          f'as it has been zipped, '
                                          f' but permission denied.'
                                          f' Try running again as admin. '
                                          f'More Info: {traceback.format_exc()}'
                                          f'\n\n'
                               )
        return 0

    def _on_run(self):
        """Command to run scrubber on target(s)"""

        # Check if target is valid and update status
        if self.input_folder == '' and self.input_file == '':
            if self._vars['Batch Size'].get() == 'Folder':
                self.status.set('No target chosen. \n'
                                'Please choose a target folder...')
            else:
                self.status.set('No target chosen. \n'
                                'Please choose a target zip file...')
        else:
            # Process zip file(s)
            if self.input_folder == '':
                self.status.set(
                    f'Processing {self.input_file} ...')
                Application.update(self)
                # Process target zip
                self.err_present = self.process_exports(self.input_file,
                                     self._vars['Post Job'].get(),
                                     self._vars['Zip?'].get(),
                                     )
                self.input_file = ''
            else:
                for file in os.listdir(self.input_folder):
                    self.status.set(f'Processing {file} ...')
                    Application.update(self)
                    # Process all zips in target directory
                    if '.zip' in file:
                        self.err_present = self.process_exports(
                            self.input_folder + '/' + file,
                            self._vars['Post Job'].get(),
                            self._vars['Zip?'].get(),
                        )
                        if self.err_present == 1:
                            self.err_count += 1
                self.input_folder = ''

            # Check and alert on errors present during run
            if self.err_count == 0 and self.err_present == 0:
                self.status.set('Processing Complete. '
                                'Add more targets to continue.')
            else:
                self.status.set('Processing Complete, '
                                'but errors are present.'
                                '\nPlease refer to the error file which will '
                                'be contained in the target directory.')
                self.err_count = 0

    def _on_target(self):
        """Command to choose a target folder/file"""

        if self._vars['Batch Size'].get() == 'Folder':
            # Request target directory
            ch_folder_diag = tk.Tk()
            ch_folder_diag.overrideredirect(True)
            ch_folder_diag.attributes('-alpha', 0)
            ch_folder_diag.title('Choose target folder...')
            self.input_folder = filedialog.askdirectory(
                title='Choose target folder...')
            ch_folder_diag.destroy()
            self.input_file = ''
        else:
            # or request target zip file
            ch_file_diag = tk.Tk()
            ch_file_diag.overrideredirect(True)
            ch_file_diag.attributes('-alpha', 0)
            ch_file_diag.title('Choose target file...')
            self.input_file = filedialog.askopenfilename(
                title='Choose target file...',
                filetypes=[("Zip Files", "*.zip")])
            ch_file_diag.destroy()
            self.input_folder = ''

        # Update status label with target info
        if self.input_folder != '':
            self.status.set(
                f'Target folder set to: \n{self.input_folder}. '
                f'\nChoose Run to continue...'
                            )
        else:
            self.status.set(f'Target file set to: \n{self.input_file}. '
                            f'\nChoose Run to continue...'
                            )

    @staticmethod
    def log_error(err_file, message) -> None:
        """Simple method for opening passed txt file
        and appending message

        :param err_file: txt file
        :param message: string to append.
        This method will prepend current date and time
        """

        with open(err_file, 'a') as f:
            f.write(f'Error_{datetime.datetime.now()}_{message}\n')

    @staticmethod
    def _on_quit():
        """Command to exit program"""
        sys.exit()


class Application(tk.Tk):
    """Application root window"""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.m_page = ''
        self.main_label = ''
        self.title(" TPG ITG Export Scrubber v1.66")
        self.minsize(400, 350)
        self.main_page()

    def main_page(self):
        self.m_page = MainPage(self)
        self.main_label = ttk.Label(
            self,
            text="TPG ITG Export Scrubber",
            font=("TKDefaultFont", 14))
        self.main_label.grid(row=0)
        self.m_page.grid(row=1, padx=10, sticky=(tk.W + tk.E))


if __name__ == "__main__":
    app = Application()
    app.grid_columnconfigure(0, weight=1)
    app.mainloop()