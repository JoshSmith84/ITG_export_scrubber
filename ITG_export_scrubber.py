"""
ITG_export_scrubber.py

Author: Josh Smith

Purpose: Stage ITG export data into a single workbook per client with
separate worksheets based on input data csv's. Format cells based on
width of widest string, freeze header row/format as a table,
and zip for sending. Also edit cell data to remove html, unicode issues,
 and leave only readable data.
 """

# imports
import os
import shutil
import csv
import unicodedata
from bs4 import BeautifulSoup
from zipfile import ZipFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import logging

# Logging config
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %'
                                                '(levelname)s - %'
                                                '(message)s'
                    )

# Logging trigger
logging.disable(logging.CRITICAL)

# Get input directory (tkinter)

# Iterate through every zip file (outer loop)

# Define global variables
working_dir = "U:\\Joshua\\Work-Stuff\\ITG stuff\\example_exports\\"
input_zip = f"{working_dir}export 1.zip"
export_dir = f"{working_dir}temp\\"
keep_csv = [
    'applications-licensing.csv', 'backup.csv', 'backups-managed.csv',
    'battery-backup-ups.csv', 'configurations.csv', 'domain-hosting.csv',
    'email.csv', 'file-sharing.csv', 'internet-wan.csv', 'lan.csv',
    'passwords.csv', 'printing.csv', 'vendors.csv',
    'voice-pbx-fax.csv', 'wireless.csv',
            ]
html_detection = ['<div>', '<br>', '<p>', '<tr>', '<tbody>', '<td>',
                  '<ol>', '<li>' '<a>', '<ul>',
                  ]
header_blue = '3498DB'
light_blue = 'D6EAF8'
h_blue_fill = PatternFill(
    start_color=header_blue, end_color=header_blue, fill_type='solid')
l_blue_fill = PatternFill(
    start_color=light_blue, end_color=light_blue, fill_type='solid')
font_header = Font(size=12)

# Unzip input
with ZipFile(input_zip, 'r') as zip:
    zip.extractall(export_dir)

# Gather list of files
input_files = []
for file in os.listdir(export_dir):
    input_files.append(file)

# Delete the deprecated backup csv if backups-managed is present
delete_backup_csv = 0
if 'backups-managed.csv' in input_files:
    delete_backup_csv += 1
logging.debug(f'original list: {input_files}\n')

# Trim the list of working csvs down to what needs to be shared
top_index_input_files = len(input_files) - 1
for index, value in enumerate(reversed(input_files)):
    if value not in keep_csv:
        del input_files[top_index_input_files - index]
        continue
    elif delete_backup_csv == 1 and value == 'backup.csv':
        del input_files[top_index_input_files - index]
        continue
logging.debug(f'edited list: {input_files}\n')

# From any of the csvs, pull customer name from column B
with open(export_dir + input_files[0], 'r', encoding='utf-8') as csv_file:
    headers = csv_file.readline().strip('\n').split(',')
    reader = csv.reader(csv_file)
    customer_name = list(reader)[0][1]
    logging.debug(f'Customer name: {customer_name}\n')

# Create output Excel workbook file based on name of the company
wb = Workbook()
wb_file = working_dir + f'{customer_name}_export.xlsx'
if os.path.exists(wb_file):
    os.remove(wb_file)
wb.save(wb_file)

# (function; inner loop)
# Iterate through every remaining csv, and make changes in memory
for file in input_files:
    # Reset columns to delete list (columns that always are deleted)
    # for each iteration,
    # so new can be added as empty columns are detected
    delete_columns = ['id', 'organization', 'Category',
                      'Business Impact', 'Client Subject Matter Expert',
                      'Importance', 'archived',
                      'Backup Estimated Start Date',
                      'FlexAssset Review Date', 'FlexAsset Review Date',
                      'Backup Radar Reporting Schedule', 'hostname',
                      'manufacturer', 'position', 'contact', 'location',
                      'configuration_interfaces', 'DHCP Exclusions',
                      'one_time_password', 'Printer Management Login',
                      'installed_by', 'Equipment make & Model',
                      'Printer Name',
                      ]

    # Continue with unpacking current csv to list of lists
    working_rows = []
    with open(export_dir + file, 'r', encoding='utf-8') as csv_file:
        headers = csv_file.readline().strip('\n').split(',')
        reader = csv.reader(csv_file)
        for row in reader:
            new_row = []
            for cell in row:
                for i in html_detection:
                    if i in cell:
                        cell = BeautifulSoup(cell, 'lxml').text
                cell = unicodedata.normalize('NFKD', cell)
                new_row.append(cell)
            working_rows.append(new_row)

    # Find the archive column and keep track of it
    # (it's usually last but not always)

    # Also find config status column
    archive_index = -1
    configuration_status_index = 0
    for index, value in enumerate(headers):
        if value == 'archived':
            archive_index = index
        if file == 'configurations.csv':
            if value == 'configuration_status':
                configuration_status_index = index
    logging.debug(f'File: {file}. Archive index #: {archive_index}\n')

    # go through every row and delete any row with archive set to 'Yes'
    # and any configuration status in configurations csv other than Active
    top_index_current = len(working_rows) - 1
    for index, value in enumerate(reversed(working_rows)):
        if value[archive_index] == 'Yes':
            del working_rows[top_index_current - index]
        if configuration_status_index != 0:
            if value[configuration_status_index] != 'Active':
                logging.debug(
                    f'Deleting File: {file}. config status index #:'
                    f' {configuration_status_index}'
                    f'Value: {value}\n')
                del working_rows[top_index_current - index]


    # Find empty columns
    for i in range(len(headers)):
        i_empty = 0
        for value in working_rows:
            if value[i] == '':
                i_empty += 1
        if i_empty == len(working_rows):
            delete_columns.append(headers[i])
    logging.debug(f'Columns to delete: {delete_columns}\n')

    # Delete all blank column index from every row
    clean_rows = []
    for row in working_rows:
        new_row = []
        for i in range(len(row)):
            if headers[i] not in delete_columns:
                new_row.append(row[i])

        clean_rows.append(new_row)

    # Clean up headers to match
    new_headers = []
    for header in headers:
        if header not in delete_columns:
            new_headers.append(header)

    # Open customer workbook and add new sheet
    wb = load_workbook(wb_file)
    sheet_name = file.split('.')[0]
    sheet = wb.create_sheet(sheet_name)

    # Output clean headers and data to Sheet
    sheet.append(new_headers)
    for row in clean_rows:
        sheet.append(row)

    # Find and set uniform column width
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        if max_length > 50:
            max_length = 50
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    # Further sheet formatting
    for cell in sheet['1:1']:
        cell.font = font_header
        # cell.fill = h_blue_fill
    sheet.freeze_panes = 'A2'
    row_count = sheet.max_row
    column_count = sheet.max_column

    # Commenting out for now, this solution is ugly once
    # data is sorted after the fact

    # for x in range(1, column_count + 1):
    #     for i in range(1, row_count + 1):
    #         c = sheet.cell(row=i, column=x)
    #         if i % 2 == 0:
    #             c.fill = l_blue_fill
    wb.save(wb_file)

# Delete the starting "Blank" sheet and delete temp files
wb = load_workbook(wb_file)
del wb['Sheet']
wb.save(wb_file)
shutil.rmtree(export_dir)

# Format sheet

# Zip the output workbook; restart outer loop